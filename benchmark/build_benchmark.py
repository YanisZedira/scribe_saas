"""Génère benchmark.xlsx (coût / performance par famille d'API).

Tarifs : juin 2026, documentation publique (voir docs/03_benchmark.md).
Usage :  python benchmark/build_benchmark.py
Puis recalcul des formules :  python <skills>/xlsx/scripts/recalc.py benchmark/benchmark.xlsx
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HDR_FILL = PatternFill("solid", start_color="1F2A55")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F2A55")
BOLD = Font(name=FONT, bold=True)
BLUE = Font(name=FONT, color="0000FF")  # entrées modifiables
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def write_table(ws, start_row, headers, rows, widths=None):
    for j, hduh in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=hduh)
    style_header(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = Font(name=FONT)
            cell.alignment = WRAP
            cell.border = BORDER
        r += 1
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
    return r


def main():
    wb = Workbook()

    # ───────────────────── Feuille 1 : Hypothèses ──────────────────────────
    ws = wb.active
    ws.title = "Hypotheses"
    ws["A1"] = "Scribe — Benchmark API (juin 2026)"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Hypothèses (modifiables — texte bleu)"
    ws["A3"].font = BOLD
    assumptions = [
        ("Durée réunion de référence (min)", 30),
        ("Réunions / utilisateur / mois", 20),
        ("Participants moyens", 3),
        ("Taux de change USD→EUR", 0.92),
        ("Tokens LLM input / réunion (k)", 7),
        ("Tokens LLM output / réunion (k)", 1),
    ]
    r = 4
    for label, val in assumptions:
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT)
        c = ws.cell(row=r, column=2, value=val)
        c.font = BLUE
        r += 1
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    # Références nommées简单 via cellules
    fx = "Hypotheses!$B$7"   # USD→EUR
    dur = "Hypotheses!$B$4"  # durée min

    # ───────────────────── Feuille 2 : Visio ───────────────────────────────
    ws = wb.create_sheet("Visio")
    ws["A1"] = "Famille 1 — Visioconférence intégrable"
    ws["A1"].font = TITLE_FONT
    headers = ["Plateforme", "Modèle", "Coût ($)", "Unité", "Récup. audio",
               "Langues", "RGPD/UE", "Verdict"]
    rows = [
        ["LiveKit", "Open-source + Cloud", 0.0005, "$/participant-min",
         "Egress multipiste ★★★", "N/A", "self-host UE ★★★", "Plateforme propre"],
        ["Recall.ai (bot)", "Teams/Meet/Zoom", 0.50, "$/h enregistrement",
         "mix + diar. ★★★", "N/A", "US ⚠️", "Visio externe"],
        ["Daily", "SaaS", 0.004, "$/participant-min (approx)",
         "raw-tracks ★★★", "N/A", "US ⚠️", "Alternative"],
        ["Jitsi", "Open-source", 0.0, "gratuit (infra)",
         "via Jibri ★★", "N/A", "self-host UE ★★★", "Repli low-cost"],
        ["Twilio Video", "SaaS", 0.0, "EOL", "—", "—", "—", "Écarté (fin de vie)"],
    ]
    end = write_table(ws, 3, headers, rows,
                      widths=[16, 20, 10, 22, 20, 8, 16, 18])
    ws.cell(row=end + 1, column=1, value="Coût Recall / réunion 30 min (€) :").font = BOLD
    # 0,5 h * coût/h * fx
    ws.cell(row=end + 1, column=3,
            value=f"=(Visio!C5)*({dur}/60)*{fx}").font = BOLD

    # ───────────────────── Feuille 3 : STT ─────────────────────────────────
    ws = wb.create_sheet("STT")
    ws["A1"] = "Famille 2 — Transcription + diarisation"
    ws["A1"].font = TITLE_FONT
    headers = ["API", "Modèle", "Coût ($/h)", "Diarisation", "Langues",
               "RGPD/UE", "Coût/réunion (€)", "Verdict"]
    rows = [
        ["OpenAI", "gpt-4o-transcribe", 0.36, "via pyannote/API", 90, "US ⚠️", None, "Socle"],
        ["AssemblyAI", "Universal", 0.27, "native (95 l.)", 99, "US ⚠️", None, "Si diar. clé"],
        ["Deepgram", "Nova-3", 0.29, "native (add-on)", 30, "US/on-prem", None, "Cible"],
        ["Whisper self-host", "large-v3 + pyannote", 0.0, "pyannote", 90, "UE ★★★", None, "Avancé"],
        ["Google STT v2", "Chirp", 0.96, "native", 125, "US ⚠️", None, "Écarté (coût)"],
    ]
    end = write_table(ws, 3, headers, rows,
                      widths=[16, 20, 11, 16, 9, 12, 16, 16])
    # Coût/réunion = coût/h * (durée/60) * fx
    for i in range(4, 4 + len(rows)):
        ws.cell(row=i, column=7, value=f"=C{i}*({dur}/60)*{fx}")
        ws.cell(row=i, column=7).number_format = "0.000 €"

    # ───────────────────── Feuille 4 : LLM ─────────────────────────────────
    ws = wb.create_sheet("LLM")
    ws["A1"] = "Famille 3 — LLM (résumé + actions)"
    ws["A1"].font = TITLE_FONT
    headers = ["Modèle", "$/1M input", "$/1M output", "Qualité FR", "JSON",
               "Coût/réunion (€)", "Verdict"]
    rows = [
        ["GPT-4o-mini", 0.15, 0.60, "★★★", "oui", None, "Défaut"],
        ["Claude Haiku", 1.00, 5.00, "★★★", "oui", None, "Repli"],
        ["Gemini 2.5 Flash", 0.30, 2.50, "★★", "oui", None, "Alternative"],
    ]
    end = write_table(ws, 3, headers, rows, widths=[18, 12, 12, 10, 8, 16, 14])
    # Coût = (in_k*price_in + out_k*price_out)/1000 * fx
    ink, outk = "Hypotheses!$B$8", "Hypotheses!$B$9"
    for i in range(4, 4 + len(rows)):
        ws.cell(row=i, column=6,
                value=f"=(({ink}*B{i})+({outk}*C{i}))/1000*{fx}")
        ws.cell(row=i, column=6).number_format = "0.0000 €"

    # ───────────────────── Feuille 5 : Synthèse coût ───────────────────────
    ws = wb.create_sheet("Synthese")
    ws["A1"] = "Synthèse — coût d'une réunion de 30 min"
    ws["A1"].font = TITLE_FONT
    headers = ["Scénario", "Captation (€)", "Transcription (€)", "LLM (€)", "Total (€)"]
    ws_rows_start = 3
    write_table(ws, ws_rows_start, headers, [
        ["Visio externe (Recall+OpenAI+GPT-4o-mini)", None, None, None, None],
        ["Plateforme propre (LiveKit+OpenAI)", None, None, None, None],
        ["Dictaphone managé (OpenAI)", None, None, None, None],
        ["Self-host UE (LiveKit+Whisper)", None, None, None, None],
    ], widths=[42, 16, 16, 12, 12])
    # Lignes 4..7
    # Captation
    ws["B4"] = f"=Visio!C5*({dur}/60)*{fx}"        # Recall 0,5 $/h
    ws["B5"] = 0
    ws["B6"] = 0
    ws["B7"] = 0
    # Transcription (Recall inclut une option, on prend OpenAI hors self-host)
    ws["C4"] = "=STT!G4"   # OpenAI €/réunion
    ws["C5"] = "=STT!G4"
    ws["C6"] = "=STT!G4"
    ws["C7"] = "=STT!G7"   # Whisper self-host ~0
    # LLM
    for rr in range(4, 8):
        ws[f"D{rr}"] = "=LLM!F4"   # GPT-4o-mini
    # Total
    for rr in range(4, 8):
        ws[f"E{rr}"] = f"=B{rr}+C{rr}+D{rr}"
        ws[f"E{rr}"].number_format = "0.000 €"
        ws[f"E{rr}"].font = BOLD
        for cc in ("B", "C", "D"):
            ws[f"{cc}{rr}"].number_format = "0.000 €"
    ws["A9"] = "Cible coût/réunion : ≤ 0,30 € — atteinte hors bot externe."
    ws["A9"].font = Font(name=FONT, italic=True, color="008000")

    out = os.path.join(os.path.dirname(__file__), "benchmark.xlsx")
    wb.save(out)
    print("écrit :", out)


if __name__ == "__main__":
    main()
